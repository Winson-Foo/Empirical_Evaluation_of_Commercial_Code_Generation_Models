#%%
'''
follow the guide of webApi by SonarCloud
https://docs.sonarcloud.io/advanced-setup/web-api/
'''
import requests
import os
import openpyxl

def CountFiles(folder_name, file_name):
    file_count = 0
    for froot_dir, cur_dir,files in os.walk(os.getcwd() + '\\' + folder_name + '\\' + file_name):
        file_count += len(files)
    return file_count

def ReadFiles():
    repo = []
    workbook = openpyxl.load_workbook('repo.xlsx')
    worksheet = workbook.active
    # iterate through the column
    for column in worksheet.iter_cols(min_col=1, max_col=1):
        for cell in column:
            repo.append(cell.value)

    return repo

# get the auth key from SonarCloud
with open ("key.txt", "r") as myfile:
    auth_token = myfile.read().splitlines()[0]

SUM_FILENAME = 'summary.xlsx'
# Create a new workbook object
HEADER_ARR = ["fileName","File Count","Issues","BUG","CODE_SMELL","VULNERABILITY","INFO","MINOR","MAJOR","CRITICAL","BLOCKER"]
FOLDER_NAME = ["Before_Refactor","After_Refactor_1","After_Refactor_2","After_Refactor_3"]
REPO_NAME = ReadFiles()

row_count = 2
# adding header and parameter
hed = {'Authorization': 'Bearer ' + auth_token}
main_url = 'https://sonarcloud.io/api/'
url = main_url + "issues/search"
workbook = openpyxl.Workbook()
# Get the active worksheet and clear its contents
worksheet = workbook.active

# adding header to the excel file
for i in range(len(HEADER_ARR)):
    row_alpha = 65 + i
    worksheet[chr(row_alpha) + "1"] = HEADER_ARR[i]

# going through the results
for i in range(len(FOLDER_NAME)):
    # 
    worksheet['A' + str(row_count)] = FOLDER_NAME[i]
    row_count = row_count + 1
    for j in range(len(REPO_NAME)):
        # variable for storring 
        issues_matrix = [0,{"BUG":0 , "CODE_SMELL":0 , "VULNERABILITY":0},{"INFO":0 , "MINOR":0 , "MAJOR":0 , "CRITICAL":0 , "BLOCKER":0}]
        param = {'componentKeys' : 'Winson-Foo_Empirical_Evaluation_of_Commercial_Code_Generation_Models:' + FOLDER_NAME[i] + '/' +REPO_NAME[j], 'resolved':'false','branch':'test','ps':'500'}
        response = requests.get(url, headers=hed, params=param)
        all_issues = response.json()['issues']
        # going through all of the issus
        for k in range(len(all_issues)):
            # number of issues
            issues_matrix[0] = issues_matrix[0] + 1
            # type of errer
            issues_matrix[1][all_issues[k]['type']] = issues_matrix[1][all_issues[k]['type']] + 1
            # severity of the error
            issues_matrix[2][all_issues[k]['severity']] = issues_matrix[2][all_issues[k]['severity']] + 1

        # writting it into the array and adding it to the excel file
        xsls_arr = []
        xsls_arr.append(REPO_NAME[j])
        xsls_arr.append(CountFiles(FOLDER_NAME[i],REPO_NAME[j]))
        xsls_arr.append(issues_matrix[0])
        type_list = list(issues_matrix[1].values())

        for k in range(len(type_list)):
            xsls_arr.append(type_list[k])

        sev_list = list(issues_matrix[2].values())

        for k in range(len(sev_list)):
            xsls_arr.append(sev_list[k])

        worksheet.append(xsls_arr)
        row_count = row_count + 1

    # adding one more row for cleanliness
    row_count = row_count + 1


# Save the changes to the XLSX file
workbook.save(SUM_FILENAME)
workbook.close()